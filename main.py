import os
import threading
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from controller.robot_controller import RobotController


APP_TITLE = "Comparador de Planilhas - Cessão vs Bases"
WINDOW_SIZE = "960x680"

OPERACOES_FILTRAR = {"DIG", "GDC", "SEM CESSAO", "CAPITAL"}

def main():
    print("Iniciando Robô...")
    controller = RobotController()
    controller.run()

if __name__ == "__main__":
    main()

def normalizar_texto(valor):
    if pd.isna(valor):
        return ""
    return str(valor).strip()


def limpar_ccb_texto(valor):
    """
    Remove traços e espaços, mantendo como texto.
    Útil apenas para fallback/log, não para o merge principal.
    """
    texto = normalizar_texto(valor)
    return texto.replace("-", "").replace(" ", "")


def limpar_ccb_para_numero(valor):
    """
    Remove traços e converte para número inteiro, sem completar dígitos.
    Exemplos:
    123-456-7   -> 1234567
    12-345-678  -> 12345678
    """
    texto = normalizar_texto(valor)
    texto = texto.replace("-", "").replace(" ", "")

    if texto == "":
        return pd.NA

    # Se vier como algo tipo "1234567.0", tenta limpar
    if texto.endswith(".0"):
        texto = texto[:-2]

    # Mantém apenas dígitos
    if texto.isdigit():
        return int(texto)

    # Se houver lixo inesperado, tenta extrair só os dígitos
    apenas_digitos = "".join(ch for ch in texto if ch.isdigit())
    if apenas_digitos:
        return int(apenas_digitos)

    return pd.NA


def detectar_separador_csv(caminho_arquivo):
    with open(caminho_arquivo, "r", encoding="utf-8-sig", errors="ignore") as f:
        primeira_linha = f.readline()

    if primeira_linha.count(";") > primeira_linha.count(","):
        return ";"
    return ","


def ler_csv_flex(caminho_arquivo, log_callback):
    separador = detectar_separador_csv(caminho_arquivo)
    log_callback(f"Lendo CSV: {os.path.basename(caminho_arquivo)} | Separador: '{separador}'")

    for encoding in ["utf-8-sig", "latin1"]:
        try:
            df = pd.read_csv(
                caminho_arquivo,
                sep=separador,
                dtype=str,
                encoding=encoding
            )
            log_callback(f"CSV lido com sucesso usando encoding: {encoding}")
            return df
        except Exception:
            continue

    raise Exception(f"Não foi possível ler o CSV: {caminho_arquivo}")


def validar_colunas(df, colunas_necessarias, nome_arquivo):
    faltantes = [col for col in colunas_necessarias if col not in df.columns]
    if faltantes:
        raise Exception(
            f"No arquivo '{nome_arquivo}' faltam as colunas obrigatórias: {', '.join(faltantes)}"
        )


def deve_filtrar(valor):
    if pd.isna(valor):
        return True

    texto = str(valor).strip()

    if texto == "":
        return True

    if texto.upper() == "#N/D":
        return True

    if texto.upper() in OPERACOES_FILTRAR:
        return True

    return False


def classificar_resumo(valor):
    if pd.isna(valor):
        return "VAZIO / NaN"

    texto = str(valor).strip()

    if texto == "":
        return "VAZIO / NaN"

    if texto.upper() == "#N/D":
        return "#N/D"

    return texto.upper()


def encontrar_coluna_por_nome(ws, nome_coluna):
    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=1, column=col).value
        if valor is not None and str(valor).strip() == nome_coluna:
            return col
    raise Exception(f"Coluna '{nome_coluna}' não encontrada na planilha Cessão.")


def copiar_estilo_celula(origem, destino):
    if origem.has_style:
        destino.font = copy(origem.font)
        destino.fill = copy(origem.fill)
        destino.border = copy(origem.border)
        destino.alignment = copy(origem.alignment)
        destino.number_format = copy(origem.number_format)
        destino.protection = copy(origem.protection)


def copiar_largura_coluna(ws_origem, ws_destino, col_origem_idx, col_destino_idx):
    letra_origem = get_column_letter(col_origem_idx)
    letra_destino = get_column_letter(col_destino_idx)
    largura = ws_origem.column_dimensions[letra_origem].width
    if largura is not None:
        ws_destino.column_dimensions[letra_destino].width = largura


def copiar_configuracoes_gerais(ws_origem, ws_destino):
    ws_destino.sheet_view.showGridLines = ws_origem.sheet_view.showGridLines
    ws_destino.freeze_panes = ws_origem.freeze_panes
    ws_destino.sheet_format.defaultRowHeight = ws_origem.sheet_format.defaultRowHeight
    ws_destino.sheet_format.defaultColWidth = ws_origem.sheet_format.defaultColWidth


def criar_planilha_filtrada_com_formatacao(
    ws_origem,
    linhas_manter,
    operacoes_filtradas,
    caminho_saida,
    log_callback=None,
    cancel_callback=None
):
    wb_novo = Workbook()
    ws_novo = wb_novo.active
    ws_novo.title = ws_origem.title

    copiar_configuracoes_gerais(ws_origem, ws_novo)

    col_contrato = encontrar_coluna_por_nome(ws_origem, "CONTRATO CRED")
    col_ccb_investidor = encontrar_coluna_por_nome(ws_origem, "CCB INVESTIDOR")
    nova_coluna_idx = col_contrato + 1

    total_colunas_novas = ws_origem.max_column + 1

    # larguras
    for col in range(1, total_colunas_novas + 1):
        if col < nova_coluna_idx:
            copiar_largura_coluna(ws_origem, ws_novo, col, col)
        elif col == nova_coluna_idx:
            copiar_largura_coluna(ws_origem, ws_novo, col_contrato, col)
        else:
            copiar_largura_coluna(ws_origem, ws_novo, col - 1, col)

    # cabeçalho
    if 1 in ws_origem.row_dimensions:
        ws_novo.row_dimensions[1].height = ws_origem.row_dimensions[1].height

    for col_nova in range(1, total_colunas_novas + 1):
        if cancel_callback:
            cancel_callback()

        if col_nova < nova_coluna_idx:
            col_origem = col_nova
            valor = ws_origem.cell(row=1, column=col_origem).value
            cel_origem = ws_origem.cell(row=1, column=col_origem)

        elif col_nova == nova_coluna_idx:
            valor = "OPERACAO FRONT"
            cel_origem = ws_origem.cell(row=1, column=col_contrato)

        else:
            col_origem = col_nova - 1
            valor = ws_origem.cell(row=1, column=col_origem).value
            cel_origem = ws_origem.cell(row=1, column=col_origem)

        cel_destino = ws_novo.cell(row=1, column=col_nova, value=valor)
        copiar_estilo_celula(cel_origem, cel_destino)

    # dados
    for idx_destino, linha_origem in enumerate(linhas_manter, start=2):
        if cancel_callback:
            cancel_callback()

        if linha_origem in ws_origem.row_dimensions:
            ws_novo.row_dimensions[idx_destino].height = ws_origem.row_dimensions[linha_origem].height

        operacao_valor = operacoes_filtradas[idx_destino - 2]

        for col_nova in range(1, total_colunas_novas + 1):
            eh_ccb_investidor = False

            if col_nova < nova_coluna_idx:
                col_origem = col_nova
                valor = ws_origem.cell(row=linha_origem, column=col_origem).value
                cel_origem = ws_origem.cell(row=linha_origem, column=col_origem)

                if col_origem == col_ccb_investidor:
                    valor = limpar_ccb_para_numero(valor)
                    eh_ccb_investidor = True

            elif col_nova == nova_coluna_idx:
                valor = operacao_valor
                cel_origem = ws_origem.cell(row=linha_origem, column=col_contrato)

            else:
                col_origem = col_nova - 1
                valor = ws_origem.cell(row=linha_origem, column=col_origem).value
                cel_origem = ws_origem.cell(row=linha_origem, column=col_origem)

                if col_origem == col_ccb_investidor:
                    valor = limpar_ccb_para_numero(valor)
                    eh_ccb_investidor = True

            cel_destino = ws_novo.cell(row=idx_destino, column=col_nova, value=valor)
            copiar_estilo_celula(cel_origem, cel_destino)

            # força formato numérico simples para não completar com zeros
            if eh_ccb_investidor:
                cel_destino.number_format = "0"

        if log_callback and ((idx_destino - 1) % 500 == 0):
            log_callback(f"Copiadas {idx_destino - 1:,} linhas para o arquivo final...")

    wb_novo.save(caminho_saida)


class App:
    def __init__(self, root):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry(WINDOW_SIZE)
        self.root.resizable(False, False)

        self.cessao_path = tk.StringVar()
        self.base1_path = tk.StringVar()
        self.base2_path = tk.StringVar()
        self.output_dir = tk.StringVar()

        self.status_var = tk.StringVar(value="Pronto.")
        self.progress_text_var = tk.StringVar(value="0%")

        self.processing = False
        self.cancel_requested = False

        self.build_ui()

    def build_ui(self):
        main = tk.Frame(self.root, padx=20, pady=20)
        main.pack(fill="both", expand=True)

        tk.Label(
            main,
            text="Comparador de Planilhas - Cessão vs Bases",
            font=("Segoe UI", 16, "bold")
        ).pack(anchor="w", pady=(0, 8))

        tk.Label(
            main,
            text="Selecione a planilha Cessão, as bases CSV e a pasta de saída.",
            font=("Segoe UI", 10)
        ).pack(anchor="w", pady=(0, 20))

        self.create_file_row(main, "Planilha Cessão (.xlsx):", self.cessao_path, self.pick_cessao)
        self.create_file_row(main, "Planilha Base 1 (.csv):", self.base1_path, self.pick_base1)
        self.create_file_row(main, "Planilha Base 2 (.csv):", self.base2_path, self.pick_base2)
        self.create_file_row(main, "Pasta de saída:", self.output_dir, self.pick_output_dir)

        btn_frame = tk.Frame(main)
        btn_frame.pack(fill="x", pady=(10, 8))

        self.btn_start = tk.Button(btn_frame, text="Iniciar", width=12, command=self.start_process)
        self.btn_start.pack(side="left", padx=(0, 8))

        self.btn_cancel = tk.Button(
            btn_frame,
            text="Cancelar",
            width=12,
            state="disabled",
            command=self.request_cancel
        )
        self.btn_cancel.pack(side="left", padx=(0, 8))

        self.btn_clear = tk.Button(btn_frame, text="Limpar logs", width=12, command=self.clear_logs)
        self.btn_clear.pack(side="left")

        tk.Label(btn_frame, textvariable=self.status_var, anchor="e").pack(side="right")

        progress_frame = tk.Frame(main)
        progress_frame.pack(fill="x", pady=(4, 14))

        self.progress = ttk.Progressbar(progress_frame, mode="determinate", maximum=100)
        self.progress.pack(side="left", fill="x", expand=True)

        tk.Label(progress_frame, textvariable=self.progress_text_var, width=6, anchor="e").pack(side="right", padx=(8, 0))

        tk.Label(main, text="Logs:", font=("Segoe UI", 10, "bold")).pack(anchor="w")

        logs_frame = tk.Frame(main)
        logs_frame.pack(fill="both", expand=True)

        self.txt_logs = tk.Text(logs_frame, height=24, wrap="word")
        self.txt_logs.pack(side="left", fill="both", expand=True)

        scrollbar = tk.Scrollbar(logs_frame, command=self.txt_logs.yview)
        scrollbar.pack(side="right", fill="y")
        self.txt_logs.config(yscrollcommand=scrollbar.set)

        self.log("Selecione os arquivos para começar.")

    def create_file_row(self, parent, label_text, var, command):
        row = tk.Frame(parent)
        row.pack(fill="x", pady=6)

        tk.Label(row, text=label_text, width=24, anchor="w").pack(side="left")
        tk.Entry(row, textvariable=var).pack(side="left", fill="x", expand=True, padx=(0, 10))
        tk.Button(row, text="Selecionar...", width=12, command=command).pack(side="right")

    def pick_cessao(self):
        path = filedialog.askopenfilename(
            title="Selecione a planilha Cessão",
            filetypes=[("Excel", "*.xlsx")]
        )
        if path:
            self.cessao_path.set(path)

    def pick_base1(self):
        path = filedialog.askopenfilename(
            title="Selecione a Base 1",
            filetypes=[("CSV", "*.csv")]
        )
        if path:
            self.base1_path.set(path)

    def pick_base2(self):
        path = filedialog.askopenfilename(
            title="Selecione a Base 2",
            filetypes=[("CSV", "*.csv")]
        )
        if path:
            self.base2_path.set(path)

    def pick_output_dir(self):
        path = filedialog.askdirectory(title="Selecione a pasta de saída")
        if path:
            self.output_dir.set(path)

    def log(self, mensagem):
        self.txt_logs.insert("end", mensagem + "\n")
        self.txt_logs.see("end")
        self.root.update_idletasks()

    def clear_logs(self):
        self.txt_logs.delete("1.0", "end")

    def update_progress(self, valor, texto=None):
        self.progress["value"] = valor
        self.progress_text_var.set(f"{int(valor)}%")
        if texto:
            self.status_var.set(texto)
        self.root.update_idletasks()

    def check_cancel(self):
        if self.cancel_requested:
            raise Exception("Processo cancelado pelo usuário.")

    def request_cancel(self):
        if self.processing:
            self.cancel_requested = True
            self.log("Cancelamento solicitado... aguardando encerramento seguro.")
            self.status_var.set("Cancelando...")

    def start_process(self):
        if self.processing:
            return

        if not self.cessao_path.get():
            messagebox.showwarning("Atenção", "Selecione a planilha Cessão.")
            return
        if not self.base1_path.get():
            messagebox.showwarning("Atenção", "Selecione a Base 1.")
            return
        if not self.base2_path.get():
            messagebox.showwarning("Atenção", "Selecione a Base 2.")
            return
        if not self.output_dir.get():
            messagebox.showwarning("Atenção", "Selecione a pasta de saída.")
            return

        self.processing = True
        self.cancel_requested = False
        self.btn_start.config(state="disabled")
        self.btn_cancel.config(state="normal")
        self.progress["value"] = 0
        self.progress_text_var.set("0%")
        self.status_var.set("Processando...")

        thread = threading.Thread(target=self.processar, daemon=True)
        thread.start()

    def finalizar_processo(self):
        self.processing = False
        self.btn_start.config(state="normal")
        self.btn_cancel.config(state="disabled")
        if self.cancel_requested:
            self.status_var.set("Cancelado.")
        elif self.status_var.get() != "Concluído com sucesso.":
            self.status_var.set("Pronto.")

    def processar(self):
        try:
            self.log("Iniciando processamento...")

            self.update_progress(5, "Lendo planilha Cessão...")
            self.check_cancel()
            df_cessao = pd.read_excel(self.cessao_path.get(), dtype=str)
            self.log(f"Cessão carregada com {len(df_cessao):,} linhas.")

            self.update_progress(15, "Lendo Base 1...")
            self.check_cancel()
            df_base1 = ler_csv_flex(self.base1_path.get(), self.log)
            self.log(f"Base 1 carregada com {len(df_base1):,} linhas.")

            self.update_progress(25, "Lendo Base 2...")
            self.check_cancel()
            df_base2 = ler_csv_flex(self.base2_path.get(), self.log)
            self.log(f"Base 2 carregada com {len(df_base2):,} linhas.")

            self.update_progress(35, "Validando colunas...")
            self.check_cancel()
            validar_colunas(df_cessao, ["CCB INVESTIDOR", "CONTRATO CRED"], "Planilha Cessão")
            validar_colunas(df_base1, ["CCB", "Operação"], "Base 1")
            validar_colunas(df_base2, ["CCB", "Operação"], "Base 2")
            self.log("Colunas obrigatórias validadas com sucesso.")

            self.update_progress(45, "Limpando CCB e convertendo para número...")
            self.check_cancel()

            # Ordem correta:
            # 1) tirar traço
            # 2) converter para número
            # 3) usar no merge
            df_cessao["CCB INVESTIDOR_NUM"] = df_cessao["CCB INVESTIDOR"].apply(limpar_ccb_para_numero)
            df_base1["CCB_NUM"] = df_base1["CCB"].apply(limpar_ccb_para_numero)
            df_base2["CCB_NUM"] = df_base2["CCB"].apply(limpar_ccb_para_numero)

            df_base1["Operação"] = df_base1["Operação"].apply(normalizar_texto)
            df_base2["Operação"] = df_base2["Operação"].apply(normalizar_texto)

            self.log("CCBs convertidos para número com sucesso.")

            self.update_progress(55, "Unindo bases...")
            self.check_cancel()

            df_bases = pd.concat(
                [df_base1[["CCB_NUM", "Operação"]], df_base2[["CCB_NUM", "Operação"]]],
                ignore_index=True
            )

            df_bases = df_bases[df_bases["CCB_NUM"].notna()].copy()
            df_bases = df_bases.drop_duplicates(subset=["CCB_NUM"], keep="first")

            self.log(f"Base consolidada com {len(df_bases):,} CCBs numéricos únicos.")

            self.update_progress(65, "Executando PROCX/MERGE numérico...")
            self.check_cancel()

            df_lookup = df_cessao.merge(
                df_bases,
                how="left",
                left_on="CCB INVESTIDOR_NUM",
                right_on="CCB_NUM"
            )

            operacao_retorno = df_lookup["Operação"].fillna("#N/D")
            self.log("Busca concluída com sucesso.")

            self.update_progress(75, "Filtrando casos desejados...")
            self.check_cancel()

            df_temp = df_cessao.copy()

            # substitui visualmente a CCB INVESTIDOR pela versão numérica limpa
            df_temp["CCB INVESTIDOR"] = df_temp["CCB INVESTIDOR"].apply(limpar_ccb_para_numero)

            idx_contrato = list(df_temp.columns).index("CONTRATO CRED")
            df_temp.insert(idx_contrato + 1, "OPERACAO FRONT", operacao_retorno.values)

            # remove coluna auxiliar se existir
            if "CCB INVESTIDOR_NUM" in df_temp.columns:
                df_temp.drop(columns=["CCB INVESTIDOR_NUM"], inplace=True)

            df_filtrado = df_temp[df_temp["OPERACAO FRONT"].apply(deve_filtrar)].copy()
            self.log(f"Total de linhas filtradas: {len(df_filtrado):,}")

            self.update_progress(85, "Montando resumo final...")
            self.check_cancel()

            resumo = df_filtrado["OPERACAO FRONT"].apply(classificar_resumo).value_counts(dropna=False)
            self.log("Resumo final por tipo de Operação:")
            for operacao, quantidade in resumo.items():
                self.log(f" - {operacao}: {quantidade:,}")

            self.update_progress(92, "Aplicando formatação e gerando arquivo final...")
            self.check_cancel()

            linhas_manter = []
            operacoes_filtradas = []

            for i, valor in enumerate(df_temp["OPERACAO FRONT"], start=2):
                if deve_filtrar(valor):
                    linhas_manter.append(i)
                    operacoes_filtradas.append(valor)

            wb_origem = load_workbook(self.cessao_path.get())
            ws_origem = wb_origem.active

            caminho_saida = os.path.join(self.output_dir.get(), "CESSAO_FILTRADA.xlsx")

            criar_planilha_filtrada_com_formatacao(
                ws_origem=ws_origem,
                linhas_manter=linhas_manter,
                operacoes_filtradas=operacoes_filtradas,
                caminho_saida=caminho_saida,
                log_callback=self.log,
                cancel_callback=self.check_cancel
            )

            self.update_progress(100, "Concluído com sucesso.")
            self.log(f"Arquivo final salvo com sucesso em: {caminho_saida}")

            messagebox.showinfo(
                "Sucesso",
                f"Arquivo gerado com sucesso:\n\n{caminho_saida}"
            )

        except Exception as e:
            self.log(f"ERRO: {str(e)}")
            if str(e) == "Processo cancelado pelo usuário.":
                messagebox.showinfo("Cancelado", "O processamento foi cancelado.")
            else:
                messagebox.showerror("Erro", str(e))
        finally:
            self.finalizar_processo()


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()